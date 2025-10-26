import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# baca data_description.txt (pastikan path-nya sesuai)
with open("data_description.txt", "r", encoding="utf-8") as f:
    text = f.read()

# ambil nama fitur sebelum tanda ':'
features_in_order = re.findall(r'(\w+):', text)
features = list(dict.fromkeys(features_in_order))  # biar unik dan urut

# kolom yang udah di-encode
encoded_cols = [
    'LotShape', 'Utilities', 'LandSlope', 'ExterQual', 'ExterCond', 
    'BsmtQual', 'BsmtCond', 'HeatingQC', 'KitchenQual', 'FireplaceQu', 
    'GarageQual', 'GarageCond', 'PoolQC', 'BsmtExposure', 'BsmtFinType1', 
    'BsmtFinType2', 'Electrical', 'Functional', 'GarageFinish', 
    'PavedDrive', 'Fence', 'Alley'
]

# bentuk jadi matriks 9x9
n = 9
matrix = [features[i:i+n] for i in range(0, len(features), n)]
if len(matrix[-1]) < n:
    matrix[-1] += [""] * (n - len(matrix[-1]))

# buat Excel baru
wb = Workbook()
ws = wb.active
ws.title = "Encoded 9x9 (Ordered)"

# warna hijau buat kolom encoded
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# isi sel dan warnai yang encoded
for r, row in enumerate(matrix, start=1):
    for c, feature in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=feature)
        if feature in encoded_cols:
            cell.fill = green_fill

# simpan file
wb.save("encoded_features_9x9_ordered.xlsx")
print("File saved as encoded_features_9x9_ordered.xlsx")
